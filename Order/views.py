from django.shortcuts import render, redirect, get_object_or_404
from .models import Order, OrderItem
from shop.models import Cart, CartItem  # adjust as needed
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_POST
import re
from django.contrib import messages
from Order.models import OrderChangeLog
from django.http import HttpResponse
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from io import BytesIO
from django.contrib.admin.views.decorators import staff_member_required
import csv
import openpyxl
from openpyxl.utils import get_column_letter


@login_required
def checkout_view(request):
    cart = getattr(request.user, 'cart', None)

    if not cart or not cart.items.exists():
        messages.warning(request, "Your cart is empty.")
        return redirect('view_cart')

    if request.method == 'POST':
        full_name = request.POST.get('full_name')
        address = request.POST.get('address')
        phone = request.POST.get('phone')

        # 🛒 Create order
        order = Order.objects.create(
            user=request.user,
            full_name=full_name,
            address=address,
            phone=phone,
            total_price=cart.total_price(),
            status='Confirmed'  # Or 'Pending'
        )

        # 🛒 Create order items
        for item in cart.items.all():
            OrderItem.objects.create(
                order=order,
                product=item.product,
                quantity=item.quantity,
                price=item.product.price
            )

        # ✅ Clear cart
        cart.items.all().delete()

        messages.success(request, "Order placed successfully!")
        return render(request, 'checkout_success.html', {'order': order})
    else:
        return render(request, 'checkout.html', {'cart': cart})

@login_required
def my_orders_view(request):
    orders = Order.objects.filter(user=request.user).order_by('-created_at')
    return render(request, 'my_orders.html', {'orders': orders})

@login_required
def order_detail(request, order_id):
    order = get_object_or_404(Order, id=order_id, user=request.user)

    if request.method == 'POST' and request.POST.get('action') == 'update_address':
        order.full_name = request.POST.get('full_name', '').strip()
        order.phone = request.POST.get('phone', '').strip()
        order.address = request.POST.get('address', '').strip()

        if not re.fullmatch(r'[6-9]\d{9}', order.phone):
            messages.error(request, "Invalid phone number. Must be 10 digits starting with 6-9.")
            return redirect('order_detail', order_id=order.id)
        
        order.save()
        OrderChangeLog.objects.create(
            order=order,
            changed_by=request.user,
            change_description=f"Updated address: {order.address}, phone: {order.phone}"
            )

        messages.success(request, "Address updated successfully.")
        return redirect('order_detail', order_id=order.id)  # Refreshes the updated info
    logs = order.change_logs.all().order_by('-changed_at')

    return render(request, 'order_detail.html', {'order': order,'logs': logs})


@require_POST
@login_required
def cancel_order(request, order_id):
    order = get_object_or_404(Order, id=order_id, user=request.user)
    if order.status != "Cancelled":
        order.status = "Cancelled"
        order.save()
    return redirect('order_detail', order_id=order.id)


@login_required
def generate_invoice(request, order_id):
    order = get_object_or_404(Order, id=order_id, user=request.user)

    buffer = BytesIO()
    p = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4

    p.setFont("Helvetica-Bold", 14)
    p.drawString(50, height - 50, f"Invoice - Order #{order.id}")
    
    p.setFont("Helvetica", 11)
    p.drawString(50, height - 80, f"Date: {order.created_at.strftime('%d %b %Y')}")
    p.drawString(50, height - 100, f"Name: {order.full_name}")
    p.drawString(50, height - 120, f"Phone: {order.phone}")
    p.drawString(50, height - 140, f"Address: {order.address}")

    y = height - 180
    p.drawString(50, y, "Products:")
    y -= 20
    for item in order.items.all():
        p.drawString(60, y, f"{item.product.name} x {item.quantity} – ₹{item.subtotal()}")
        y -= 20

    y -= 10
    p.setFont("Helvetica-Bold", 12)
    p.drawString(50, y, f"Total Amount: ₹{order.total_price}")

    p.showPage()
    p.save()

    buffer.seek(0)
    return HttpResponse(buffer, content_type='application/pdf')

from datetime import datetime

@staff_member_required
def export_orders_csv(request):
    start = request.GET.get('start')
    end = request.GET.get('end')

    orders = Order.objects.all()

    if start and end:
        try:
            start_date = datetime.strptime(start, '%Y-%m-%d')
            end_date = datetime.strptime(end, '%Y-%m-%d')
            orders = orders.filter(created_at__range=(start_date, end_date))
        except ValueError:
            pass  # skip filter if format is wrong

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="order_items.csv"'

    writer = csv.writer(response)
    writer.writerow(['Order ID', 'Customer', 'Product', 'Qty', 'Price', 'Subtotal', 'Status', 'Order Date'])

    for order in orders.prefetch_related('items', 'user'):
        for item in order.items.all():
            writer.writerow([
                order.id,
                order.user.username,
                item.product.name,
                item.quantity,
                item.price,
                item.subtotal(),
                order.status,
                order.created_at.strftime('%Y-%m-%d %H:%M')
            ])

    return response


@staff_member_required
def export_orders_excel(request):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Orders'

    headers = ['Order ID', 'Customer', 'Product', 'Qty', 'Price', 'Subtotal', 'Status', 'Order Date']
    sheet.append(headers)

    for order in Order.objects.all().prefetch_related('items', 'user'):
        for item in order.items.all():
            sheet.append([
                order.id,
                order.user.username,
                item.product.name,
                item.quantity,
                item.price,
                item.subtotal(),
                order.status,
                order.created_at.strftime('%Y-%m-%d %H:%M')
            ])

    for col in range(1, len(headers) + 1):
        sheet.column_dimensions[get_column_letter(col)].width = 15

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=orders.xlsx'
    workbook.save(response)
    return response